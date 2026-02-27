"""
CensusIQ - Employee Census Generator
Triton Benefits & HR Solutions
Run: python app.py  → opens at http://localhost:5050
"""

import os, json, re, threading, webbrowser, time
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber

# OCR support for image-based PDFs (carrier invoices)
try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = Path('uploads')
app.config['EXPORT_FOLDER'] = Path('exports')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# ─────────────────────────────────────────────
# SESSION STATE  (in-memory for desktop use)
# ─────────────────────────────────────────────
session_data = {
    "files": [],              # list of {name, type, path, status}
    "records": [],            # merged/processed employee records
    "issues": [],             # validation flags
    "merge_log": [],          # what came from where
    "company_name": "",       # group/client name (user-set)
    "detected_company": "",   # auto-detected from invoice
    "invoice_meta": {},       # invoice metadata (period, number, etc.)
    "invoice_records": [],    # raw records sourced from carrier invoices
    "census_records": [],     # raw records sourced from spreadsheets/CSVs
    "reconcile": [],          # computed reconcile rows
}

STANDARD_FIELDS = [
    "last_name", "first_name", "relationship", "gender",
    "dob", "state", "zip", "plan_election", "emp_status",
    "hire_date", "term_date", "ssn", "email", "phone",
    "salary", "hours_per_week", "waive_reason"
]

RELATIONSHIP_MAP = {
    "subscriber": "Subscriber", "employee": "Subscriber", "ee": "Subscriber",
    "self": "Subscriber", "primary": "Subscriber",
    "spouse": "Spouse", "sp": "Spouse", "wife": "Spouse", "husband": "Spouse", "partner": "Spouse",
    "child": "Child", "ch": "Child", "dependent": "Child", "dep": "Child",
    "son": "Child", "daughter": "Child",
}

GENDER_MAP = {
    "m": "M", "male": "M", "man": "M",
    "f": "F", "female": "F", "woman": "F",
}

COVERAGE_MAP = {
    "single": "Single", "ee only": "Single", "employee only": "Single", "s": "Single",
    "2 adult": "2 Adult", "ee+sp": "2 Adult", "employee+spouse": "2 Adult", "ee spouse": "2 Adult",
    "parent/ch": "Parent/CH", "ee+ch": "Parent/CH", "employee+child": "Parent/CH", "ee child": "Parent/CH",
    "family": "Fam", "fam": "Fam", "ee+fam": "Fam", "employee+family": "Fam",
    "wc": "WC - Waive with other coverage", "waive other": "WC - Waive with other coverage",
    "wo": "WO - Waive coverage (no other insurance)", "waive": "WO - Waive coverage (no other insurance)",
    "wp": "WP - Waiting Period (not yet eligible)", "waiting": "WP - Waiting Period (not yet eligible)",
}

STATUS_MAP = {
    "f": "F", "full": "F", "full-time": "F", "fulltime": "F", "ft": "F",
    "p": "P", "part": "P", "part-time": "P", "parttime": "P", "pt": "P",
    "c": "C", "cobra": "C",
    "r": "R", "retiree": "R", "retired": "R",
    "s": "S", "seasonal": "S",
}

# ─────────────────────────────────────────────
# PARSERS
# ─────────────────────────────────────────────

def parse_date(val):
    """Try many date formats, return MM/DD/YYYY string or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%Y")
    s = str(val).strip()
    if not s or s.lower() in ("none", "nan", "n/a", ""):
        return None
    formats = [
        "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%m/%d/%y",
        "%m-%d-%y", "%B %d, %Y", "%b %d, %Y", "%d-%b-%Y",
        "%d/%m/%Y", "%Y%m%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).strftime("%m/%d/%Y")
        except:
            pass
    # Try extracting numbers
    nums = re.findall(r'\d+', s)
    if len(nums) == 3:
        m, d, y = nums
        if len(y) == 2:
            y = "20" + y if int(y) < 50 else "19" + y
        try:
            return datetime(int(y), int(m), int(d)).strftime("%m/%d/%Y")
        except:
            pass
    return s  # return raw if can't parse

def normalize_ssn(val):
    if not val:
        return None
    s = re.sub(r'\D', '', str(val))
    if len(s) == 9:
        return f"{s[:3]}-{s[3:5]}-{s[5:]}"
    return str(val).strip() if str(val).strip() else None

def normalize_zip(val):
    if not val:
        return None
    s = re.sub(r'\D', '', str(val))
    if len(s) >= 5:
        return s[:5]
    return str(val).strip()

def guess_column_mapping(columns):
    """Fuzzy-match raw column names to standard fields."""
    mapping = {}
    col_patterns = {
        "last_name":    r"last|lname|surname|family.?name",
        "first_name":   r"first|fname|given.?name|forename",
        "middle_name":  r"middle|mname|mi\b",
        "relationship": r"relat|rel\b|dep.?type|member.?type",
        "gender":       r"gender|sex\b|m/?f",
        "dob":          r"birth|dob|born|date.?of.?birth",
        "state":        r"\bstate\b|residential.?state|home.?state",
        "zip":          r"zip|postal|postcode",
        "plan_election":r"plan|election|coverage|tier|benefit",
        "emp_status":   r"status|employment.?status|emp.?stat|work.?status",
        "hire_date":    r"hire|start.?date|begin.?date|original.?hire",
        "term_date":    r"term|end.?date|termination",
        "ssn":          r"ssn|social|tax.?id|tin\b|ss#|ss.?number",
        "email":        r"email|e-mail",
        "phone":        r"phone|mobile|cell|telephone",
        "salary":       r"salary|wage|pay|compensation|annual",
        "hours_per_week": r"hours|hrs.?per|weekly.?hours",
        "waive_reason": r"waive|waiver.?reason",
        "full_name":    r"full.?name|name\b",
    }
    for col in columns:
        col_clean = str(col).lower().strip()
        for field, pattern in col_patterns.items():
            if re.search(pattern, col_clean, re.IGNORECASE):
                if field not in mapping.values():
                    mapping[col] = field
                    break
    return mapping

def parse_excel_csv(filepath, filename):
    """Parse Excel or CSV file into list of raw row dicts."""
    records = []
    log = []
    try:
        ext = Path(filename).suffix.lower()
        if ext == '.csv':
            df = pd.read_csv(filepath, dtype=str)
        else:
            # Try to find the header row (skip legend rows)
            df_probe = pd.read_excel(filepath, header=None, dtype=str, nrows=30)
            header_row = 0
            for idx, row in df_probe.iterrows():
                non_null = row.dropna()
                # Look for row that looks like headers (contains name-like words)
                row_text = ' '.join(non_null.values).lower()
                if any(w in row_text for w in ['last', 'first', 'name', 'date', 'gender', 'birth', 'plan']):
                    header_row = idx
                    break
            df = pd.read_excel(filepath, header=header_row, dtype=str)

        # Drop completely empty rows
        df = df.dropna(how='all')
        # Drop rows that are all the same value (legend rows)
        df = df[df.apply(lambda r: r.dropna().nunique() > 1, axis=1)]

        col_map = guess_column_mapping(df.columns.tolist())
        log.append(f"Detected {len(df.columns)} columns, mapped {len(col_map)}")

        for _, row in df.iterrows():
            rec = {}
            for raw_col, std_field in col_map.items():
                val = row.get(raw_col)
                if pd.notna(val) and str(val).strip() not in ('', 'nan', 'None'):
                    rec[std_field] = str(val).strip()
            # Handle full_name split
            if 'full_name' in rec and 'first_name' not in rec:
                parts = rec['full_name'].split()
                if len(parts) >= 2:
                    rec['last_name'] = parts[-1]
                    rec['first_name'] = ' '.join(parts[:-1])
                del rec['full_name']

            if rec:
                rec['_source'] = filename
                records.append(rec)

    except Exception as e:
        log.append(f"ERROR parsing {filename}: {e}")

    return records, log

def detect_carrier_invoice(text):
    """Detect if a PDF is a carrier invoice and which carrier."""
    text_lower = text.lower()
    if 'unitedhealthcare' in text_lower or 'uhcservices' in text_lower or 'libhe' in text_lower or 'libhp' in text_lower:
        return 'UHC'
    if 'aetna' in text_lower:
        return 'AETNA'
    if 'cigna' in text_lower:
        return 'CIGNA'
    if 'bcbs' in text_lower or 'blue cross' in text_lower or 'blue shield' in text_lower:
        return 'BCBS'
    return None

def parse_uhc_invoice_ocr(filepath, filename):
    """Parse UHC carrier invoice using OCR. Returns (records, company_name, invoice_meta, log)."""
    if not OCR_AVAILABLE:
        return [], None, {}, ["OCR not available — install pdf2image and pytesseract"]

    UHC_COVERAGE_MAP = {
        'E': 'Single',
        'ES': '2 Adult',
        'ESC': 'Fam',
        'EC': 'Parent/CH',
        'E1D': 'Parent/CH',
        'F': 'Fam',
    }

    log = []
    employees = {}
    company_name = None
    invoice_meta = {}

    try:
        pages_img = convert_from_path(str(filepath), dpi=200)
        log.append(f"OCR: processing {len(pages_img)} pages")

        for page_num, page_img in enumerate(pages_img):
            text = pytesseract.image_to_string(page_img)
            lines = text.split('\n')

            # Extract company name from detail pages (page 3+)
            # In UHC invoices: company name appears on first line of detail pages
            # Pattern: "APM Hexseal    Page X of Y" or "APM Hexseal" alone
            if page_num >= 2 and not company_name:
                first_line = next((l.strip() for l in lines if l.strip()), '')
                # Remove "Page X of Y" suffix if present
                clean = re.sub(r'\s+Page\s+\d+\s+of\s+\d+.*', '', first_line).strip()
                if clean and len(clean) > 2 and re.match(r'^[A-Za-z]', clean) and 'Customer' not in clean:
                    company_name = clean.title()
                    log.append(f"Company name from page {page_num+1}: {company_name}")

            # Extract invoice metadata
            for line in lines:
                m = re.search(r'Coverage Period:\s*([\d/]+\s*-\s*[\d/]+)', line)
                if m: invoice_meta['coverage_period'] = m.group(1).strip()
                m = re.search(r'Invoice No[:\s]+([\d]+)', line)
                if m: invoice_meta['invoice_no'] = m.group(1).strip()
                m = re.search(r'Invoice Date[:\s]+([\d/]+)', line)
                if m: invoice_meta['invoice_date'] = m.group(1).strip()
                m = re.search(r'Customer No[:\s]+([\d]+)', line)
                if m: invoice_meta['customer_no'] = m.group(1).strip()
                m = re.search(r'Bill Group[:\s\w]*([\d]+)', line)
                if m: invoice_meta['bill_group'] = m.group(1).strip()

            # Parse employee rows - two patterns:
            # Pattern A: line with Total amount at end (first of two lines per employee)
            # Pattern B: line without Total (second component line)
            row_pattern = r'\d{6,7}\s*[\|l]\s*([A-Z][a-zA-Z\s\-,\.]+?)\s+Lib[A-Z0-9].*?[\|\|l\s]([A-Z]{1,3})\s+A'
            
            for line in lines:
                m = re.search(row_pattern, line)
                if m:
                    name_raw = m.group(1).strip().rstrip(',').strip()
                    coverage_code = m.group(2).strip()
                    if coverage_code not in ('E','ES','ESC','EC','E1D','EID','F','TE'):
                        continue

                    # Parse "Last, First" format
                    if ',' in name_raw:
                        parts = name_raw.split(',', 1)
                        last = parts[0].strip().title()
                        first = parts[1].strip().title()
                    else:
                        words = name_raw.split()
                        first = words[0].title() if words else ''
                        last = ' '.join(words[1:]).title() if len(words) > 1 else ''

                    if not last or not first:
                        continue

                    # Extract premium total (last dollar amount on line if 2+ amounts present)
                    dollars = re.findall(r'\$([\d,]+\.\d{2})', line)
                    premium_total = None
                    if len(dollars) >= 2:
                        premium_total = f"${dollars[-1]}"

                    key = f"{last.lower()}|{first.lower()}"
                    if key not in employees:
                        employees[key] = {
                            'last_name': last,
                            'first_name': first,
                            'relationship': 'Subscriber',
                            'emp_status': 'F',
                            'plan_election': UHC_COVERAGE_MAP.get(coverage_code, 'Single'),
                            '_coverage_code': coverage_code,
                            '_premium_total': premium_total,
                            '_source': filename,
                            '_source_type': 'invoice',
                            '_carrier': 'UHC',
                            '_invoice_period': invoice_meta.get('coverage_period', ''),
                        }
                    elif premium_total and not employees[key].get('_premium_total'):
                        employees[key]['_premium_total'] = premium_total

        log.append(f"OCR extracted {len(employees)} unique employees from UHC invoice")
        if company_name:
            log.append(f"Company detected: {company_name}")
        if invoice_meta:
            log.append(f"Invoice: {invoice_meta.get('invoice_no','')} | Period: {invoice_meta.get('coverage_period','')}")

    except Exception as e:
        log.append(f"OCR ERROR: {e}")
        import traceback
        log.append(traceback.format_exc()[:300])

    return list(employees.values()), company_name, invoice_meta, log


def parse_pdf(filepath, filename):
    """Extract employee data from PDF invoices/reports."""
    records = []
    log = []
    raw_text = ""

    try:
        # First check if it's an image-based PDF (carrier invoice)
        with pdfplumber.open(filepath) as pdf:
            total_chars = sum(len(p.chars) for p in pdf.pages[:3])
            is_image_pdf = total_chars == 0

            if not is_image_pdf:
                # Text-based PDF — use pdfplumber
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if not table:
                            continue
                        headers = None
                        data_rows = []
                        for i, row in enumerate(table):
                            row_text = ' '.join(str(c or '') for c in row).lower()
                            if any(w in row_text for w in ['name', 'last', 'first', 'birth', 'dob', 'gender', 'plan']):
                                headers = [str(c or '').strip() for c in row]
                                data_rows = table[i+1:]
                                break
                        if headers and data_rows:
                            col_map = guess_column_mapping(headers)
                            log.append(f"PDF table: {len(headers)} cols, {len(data_rows)} rows")
                            for row in data_rows:
                                rec = {}
                                for j, h in enumerate(headers):
                                    if j < len(row) and h in col_map:
                                        val = str(row[j] or '').strip()
                                        if val and val.lower() not in ('none', 'nan', ''):
                                            rec[col_map[h]] = val
                                if rec:
                                    rec['_source'] = filename
                                    records.append(rec)
                    raw_text += (page.extract_text() or "") + "\n"

                # Detect carrier type from text
                carrier = detect_carrier_invoice(raw_text)
                if carrier:
                    log.append(f"Detected {carrier} invoice format (text-based)")

                if not records and raw_text:
                    log.append("No tables found, trying text pattern extraction")
                    records, extra_log = extract_from_text(raw_text, filename)
                    log.extend(extra_log)

            else:
                # Image-based PDF — use OCR
                log.append("Image-based PDF detected — using OCR")

                # Quick OCR of first page to detect carrier
                if OCR_AVAILABLE:
                    try:
                        first_page_imgs = convert_from_path(str(filepath), dpi=150, first_page=1, last_page=1)
                        first_text = pytesseract.image_to_string(first_page_imgs[0])
                        carrier = detect_carrier_invoice(first_text)
                        log.append(f"OCR carrier detection: {carrier or 'Unknown'}")
                    except:
                        carrier = None

                    if carrier == 'UHC':
                        records, detected_company, invoice_meta, ocr_log = parse_uhc_invoice_ocr(filepath, filename)
                        log.extend(ocr_log)
                        # Store invoice metadata on records
                        for r in records:
                            r['_invoice_meta'] = invoice_meta
                            if detected_company and not session_data.get('company_name'):
                                session_data['detected_company'] = detected_company
                    else:
                        # Generic OCR extraction
                        log.append(f"Using generic OCR extraction for {carrier or 'unknown'} invoice")
                        pages_img = convert_from_path(str(filepath), dpi=200)
                        full_text = '\n'.join(pytesseract.image_to_string(p) for p in pages_img)
                        records, extra_log = extract_from_text(full_text, filename)
                        log.extend(extra_log)
                else:
                    log.append("WARNING: OCR not available. Install pdf2image + pytesseract to parse image PDFs.")
                    log.append("Run: pip install pdf2image pytesseract")

    except Exception as e:
        log.append(f"ERROR parsing PDF {filename}: {e}")

    return records, log

def extract_from_text(text, filename):
    """Pattern-match employee data from raw PDF text."""
    records = []
    log = []
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    # Common invoice/report patterns
    # Try to find name + DOB combos
    date_pattern = r'\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b'
    ssn_pattern = r'\b\d{3}[-\s]\d{2}[-\s]\d{4}\b'
    name_pattern = r'^([A-Z][a-z]+),?\s+([A-Z][a-z]+)'

    current_rec = {}
    for line in lines:
        # SSN
        ssn_match = re.search(ssn_pattern, line)
        if ssn_match:
            current_rec['ssn'] = ssn_match.group()

        # Dates
        dates = re.findall(date_pattern, line)
        if dates and 'dob' not in current_rec:
            current_rec['dob'] = dates[0]

        # Name
        name_match = re.match(name_pattern, line)
        if name_match:
            if current_rec and ('last_name' in current_rec or 'ssn' in current_rec):
                current_rec['_source'] = filename
                records.append(current_rec)
                current_rec = {}
            current_rec['last_name'] = name_match.group(1)
            current_rec['first_name'] = name_match.group(2)

        # Gender
        if re.search(r'\b(male|female|M|F)\b', line, re.IGNORECASE):
            g = re.search(r'\b(male|female|M|F)\b', line, re.IGNORECASE).group(1)
            current_rec['gender'] = GENDER_MAP.get(g.lower(), g.upper()[:1])

    if current_rec:
        current_rec['_source'] = filename
        records.append(current_rec)

    log.append(f"Text extraction: {len(records)} potential records found")
    return records, log

# ─────────────────────────────────────────────
# NORMALIZATION & MERGING
# ─────────────────────────────────────────────

def normalize_record(rec):
    """Normalize all fields to standard values."""
    out = dict(rec)

    if 'dob' in out:
        out['dob'] = parse_date(out['dob'])
    if 'hire_date' in out:
        out['hire_date'] = parse_date(out['hire_date'])
    if 'term_date' in out:
        out['term_date'] = parse_date(out['term_date'])

    if 'ssn' in out:
        out['ssn'] = normalize_ssn(out['ssn'])

    if 'zip' in out:
        out['zip'] = normalize_zip(out['zip'])

    if 'gender' in out:
        out['gender'] = GENDER_MAP.get(out['gender'].lower().strip(), out['gender'].upper()[:1])

    if 'relationship' in out:
        out['relationship'] = RELATIONSHIP_MAP.get(out['relationship'].lower().strip(), out['relationship'])

    if 'plan_election' in out:
        out['plan_election'] = COVERAGE_MAP.get(out['plan_election'].lower().strip(), out['plan_election'])

    if 'emp_status' in out:
        out['emp_status'] = STATUS_MAP.get(out['emp_status'].lower().strip(), out['emp_status'])

    if 'state' in out:
        out['state'] = out['state'].upper().strip()[:2]

    # Capitalize names
    for field in ('last_name', 'first_name'):
        if field in out:
            out[field] = str(out[field]).strip().title()

    return out

def make_record_key(rec):
    """Create a matching key for deduplication/merging."""
    last = (rec.get('last_name') or '').lower().strip()
    first = (rec.get('first_name') or '').lower().strip()
    dob = rec.get('dob') or ''
    ssn_last4 = ''
    if rec.get('ssn'):
        ssn_last4 = re.sub(r'\D', '', str(rec['ssn']))[-4:]
    return f"{last}|{first}|{dob}|{ssn_last4}"

def merge_records(all_records):
    """Smart merge: combine data from multiple sources for same person."""
    merged = {}
    merge_log = []

    for rec in all_records:
        key = make_record_key(rec)
        if not key.replace('|', '').strip():
            continue

        if key not in merged:
            merged[key] = dict(rec)
            merge_log.append({
                "action": "new",
                "name": f"{rec.get('first_name','')} {rec.get('last_name','')}".strip(),
                "source": rec.get('_source', ''),
                "key": key
            })
        else:
            # Merge: fill in missing fields from new source
            existing = merged[key]
            filled = []
            for field, val in rec.items():
                if field.startswith('_'):
                    continue
                if not existing.get(field) and val:
                    existing[field] = val
                    filled.append(field)
            if filled:
                existing['_sources'] = existing.get('_sources', [existing.get('_source', '')])
                if rec.get('_source') not in existing['_sources']:
                    existing['_sources'].append(rec.get('_source', ''))
                merge_log.append({
                    "action": "merged",
                    "name": f"{rec.get('first_name','')} {rec.get('last_name','')}".strip(),
                    "source": rec.get('_source', ''),
                    "filled": filled
                })

    return list(merged.values()), merge_log

def validate_records(records):
    """Flag issues in merged records."""
    issues = []
    ssns_seen = {}

    for i, rec in enumerate(records):
        row_id = f"{rec.get('first_name','')} {rec.get('last_name','')}".strip() or f"Row {i+1}"

        # Missing critical fields
        if not rec.get('last_name'):
            issues.append({"severity":"error","employee":row_id,"field":"last_name","msg":"Missing last name","row":i})
        if not rec.get('first_name'):
            issues.append({"severity":"error","employee":row_id,"field":"first_name","msg":"Missing first name","row":i})
        if not rec.get('dob'):
            issues.append({"severity":"error","employee":row_id,"field":"dob","msg":"Missing date of birth","row":i})
        if not rec.get('gender'):
            issues.append({"severity":"warning","employee":row_id,"field":"gender","msg":"Missing gender","row":i})
        if not rec.get('relationship'):
            issues.append({"severity":"warning","employee":row_id,"field":"relationship","msg":"Missing relationship","row":i})
        if not rec.get('state'):
            issues.append({"severity":"warning","employee":row_id,"field":"state","msg":"Missing residential state","row":i})
        if not rec.get('zip'):
            issues.append({"severity":"warning","employee":row_id,"field":"zip","msg":"Missing zip code","row":i})
        if not rec.get('plan_election'):
            if rec.get('relationship') in ('Subscriber', None):
                issues.append({"severity":"warning","employee":row_id,"field":"plan_election","msg":"Missing plan election (subscriber)","row":i})

        # SSN duplicate check
        ssn = rec.get('ssn')
        if ssn:
            clean_ssn = re.sub(r'\D', '', ssn)
            if clean_ssn in ssns_seen:
                issues.append({"severity":"error","employee":row_id,"field":"ssn",
                                "msg":f"Duplicate SSN — also on {ssns_seen[clean_ssn]}","row":i})
            else:
                ssns_seen[clean_ssn] = row_id
        else:
            if rec.get('relationship') == 'Subscriber':
                issues.append({"severity":"warning","employee":row_id,"field":"ssn","msg":"Missing SSN (subscriber)","row":i})

        # Date validation
        for date_field in ('dob', 'hire_date', 'term_date'):
            val = rec.get(date_field)
            if val:
                try:
                    d = datetime.strptime(val, "%m/%d/%Y")
                    if date_field == 'dob' and d.year > 2020:
                        issues.append({"severity":"info","employee":row_id,"field":date_field,
                                       "msg":f"Very recent birth year {d.year} — confirm dependent age","row":i})
                    if date_field == 'hire_date' and d > datetime.now():
                        issues.append({"severity":"warning","employee":row_id,"field":date_field,
                                       "msg":"Future hire date — verify this is correct","row":i})
                except:
                    issues.append({"severity":"warning","employee":row_id,"field":date_field,
                                   "msg":f"Unrecognized date format: {val}","row":i})

    return issues

# ─────────────────────────────────────────────
# EXPORT BUILDERS
# ─────────────────────────────────────────────

def build_triton_census(records, company_name=""):
    """Build the Triton Benefits standard census Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Census"

    # Colors
    navy = "0D1B2A"
    teal = "00C9A7"
    amber = "F5A623"
    light_gray = "F4F6F9"
    mid_gray = "E2E8F0"

    # Legend section (rows 1-18) — matches Triton template
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "EMPLOYEE CENSUS"
    title_cell.font = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=navy)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    # Coverage legend
    ws['F1'] = "Coverage Electing"
    ws['F1'].font = Font(bold=True)
    for i, label in enumerate(["Single","2 Adult","Parent/CH","Fam",
                                 "WC - Waive with other coverage",
                                 "WO - Waive coverage (no other insurance)",
                                 "WP - Waiting Period (not yet eligible)"], start=2):
        ws[f'F{i}'] = label

    ws['A9'] = "Company Name:"
    ws['B9'] = company_name or "________________________________"
    ws['F9'] = "Relationship Code"
    ws['F9'].font = Font(bold=True)
    ws['A10'] = "Street Address:"
    ws['B10'] = "________________________________"
    ws['F10'] = "Subscriber"
    ws['A11'] = "City, State, Zip"
    ws['B11'] = "________________________________"
    ws['F11'] = "Spouse"
    ws['A12'] = "Business Type"
    ws['B12'] = "________________________________"
    ws['F12'] = "Child"
    ws['F13'] = "Employment Status"
    ws['F13'].font = Font(bold=True)
    for i, label in enumerate(["F-Full-time","P-Part-Time","C-COBRA","R-Retiree","S-Seasonal"], start=14):
        ws[f'F{i}'] = label

    ws['A20'] = "EXAMPLE:"
    ws['A20'].font = Font(bold=True, color=navy)

    # Example rows
    ws.append(["Employee Last Name", "Employee First Name", "Relationship", "Gender",
                "Date of Birth", "Residential State", "Residential Zip Code", "Plan Election", "Emp Status"])
    for cell in ws[21]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal='center')

    ws.append(["Smith", "John", "Subscriber", "M", "06/22/1982", "NJ", "07095", "PPO 1500", "F"])
    ws.append(["Smith", "Jane", "Spouse", "F", "06/23/1981", "NJ", "07095", "", ""])
    ws.append(["Smith", "Little Johnny", "Child", "M", "06/23/2020", "NJ", "07095", "", ""])
    ws.append([])

    # Column widths
    col_widths = [20, 20, 14, 8, 14, 8, 12, 18, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # DATA HEADER ROW
    header_row = ws.max_row + 1
    headers = ["Employee Last Name", "Employee First Name", "Relationship", "Gender",
               "Date of Birth", "Residential State", "Residential Zip Code", "Plan Election", "Emp Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1A2D42")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[header_row].height = 20

    # DATA ROWS
    alt = False
    for rec in records:
        data_row = [
            rec.get('last_name', ''),
            rec.get('first_name', ''),
            rec.get('relationship', ''),
            rec.get('gender', ''),
            rec.get('dob', ''),
            rec.get('state', ''),
            rec.get('zip', ''),
            rec.get('plan_election', ''),
            rec.get('emp_status', ''),
        ]
        row_num = ws.max_row + 1
        for col, val in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=light_gray if alt else "FFFFFF")
            cell.alignment = Alignment(horizontal='center' if col > 2 else 'left', vertical='center')
            cell.font = Font(size=10)
        alt = not alt

    # Freeze panes at data
    ws.freeze_panes = ws.cell(row=header_row+1, column=1)

    return wb

def build_carrier_file(records, carrier):
    """Build carrier-specific formatted census."""
    carrier_schemas = {
        "BCBS": {
            "fields": ["last_name","first_name","dob","gender","ssn","relationship",
                       "state","zip","plan_election","hire_date","emp_status"],
            "headers": ["Last Name","First Name","Date of Birth","Gender","SSN","Relationship",
                        "State","Zip","Plan","Hire Date","Status"],
            "color": "003087"
        },
        "Aetna": {
            "fields": ["last_name","first_name","relationship","dob","gender",
                       "ssn","zip","state","plan_election","emp_status","hire_date"],
            "headers": ["Last Name","First Name","Rel","DOB","Sex","SSN",
                        "Zip","State","Coverage Tier","Status","Hire Date"],
            "color": "7B0C2A"
        },
        "UHC": {
            "fields": ["last_name","first_name","gender","dob","ssn",
                       "relationship","state","zip","plan_election","hire_date","emp_status"],
            "headers": ["Subscriber Last","Subscriber First","Gender","DOB","Tax ID",
                        "Member Type","State","Zip Code","Plan Code","Original Hire Date","Status"],
            "color": "003366"
        },
        "Cigna": {
            "fields": ["last_name","first_name","dob","ssn","gender",
                       "relationship","zip","state","plan_election","hire_date"],
            "headers": ["Last Name","First Name","Birth Date","SSN","Gender",
                        "Member Type","Zip","State","Benefit Plan","Hire Date"],
            "color": "006B50"
        },
    }

    schema = carrier_schemas.get(carrier, carrier_schemas["BCBS"])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{carrier} Census"

    # Header
    color = schema["color"]
    for col, h in enumerate(schema["headers"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.row_dimensions[1].height = 18

    for i, rec in enumerate(records, 2):
        for col, field in enumerate(schema["fields"], 1):
            ws.cell(row=i, column=col, value=rec.get(field, ''))
            if i % 2 == 0:
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="F8F9FA")

    ws.freeze_panes = 'A2'
    return wb


# ─────────────────────────────────────────────
# RECONCILE ENGINE
# ─────────────────────────────────────────────

TIER_NORMALIZE = {
    'single': 'Single', 'employee only': 'Single', 'ee only': 'Single',
    'ee': 'Single', 'e': 'Single',
    '2 adult': '2 Adult', 'ee+sp': '2 Adult', 'employee+spouse': '2 Adult',
    'employee spouse': '2 Adult', 'ee spouse': '2 Adult', 'es': '2 Adult',
    'parent/ch': 'Parent/CH', 'ee+ch': 'Parent/CH', 'employee+child': 'Parent/CH',
    'parent child': 'Parent/CH', 'ec': 'Parent/CH', 'e1d': 'Parent/CH',
    'fam': 'Fam', 'family': 'Fam', 'ee+fam': 'Fam', 'employee+family': 'Fam',
    'employee family': 'Fam', 'esc': 'Fam',
    'wc - waive with other coverage': 'Waived', 'wc': 'Waived',
    'wo - waive coverage (no other insurance)': 'Waived', 'wo': 'Waived',
    'wp - waiting period (not yet eligible)': 'Waiting', 'wp': 'Waiting',
}

def norm_tier(val):
    if not val: return None
    return TIER_NORMALIZE.get(str(val).lower().strip(), str(val).strip())

def fuzzy_name_key(last, first):
    """Normalize name for matching — handles OCR typos and case differences."""
    def clean(s):
        return re.sub(r'[^a-z]', '', str(s or '').lower())
    return clean(last)[:6] + '|' + clean(first)[:4]

def build_reconcile(all_records):
    """
    Compare invoice-sourced records vs census-sourced records.
    Returns list of reconcile rows with match status and discrepancies.
    """
    invoice_recs = [r for r in all_records if r.get('_source_type') == 'invoice']
    census_recs  = [r for r in all_records if r.get('_source_type') != 'invoice']

    # Build lookup dicts by fuzzy key
    invoice_by_key = {}
    for r in invoice_recs:
        k = fuzzy_name_key(r.get('last_name'), r.get('first_name'))
        if k not in invoice_by_key:
            invoice_by_key[k] = r

    census_by_key = {}
    for r in census_recs:
        if r.get('relationship') in ('Subscriber', None, ''):
            k = fuzzy_name_key(r.get('last_name'), r.get('first_name'))
            if k not in census_by_key:
                census_by_key[k] = r

    all_keys = set(list(invoice_by_key.keys()) + list(census_by_key.keys()))
    rows = []

    for key in all_keys:
        inv = invoice_by_key.get(key)
        cen = census_by_key.get(key)
        discrepancies = []
        match_status = 'matched'

        # Determine display name
        if inv:
            display_last = inv.get('last_name', '')
            display_first = inv.get('first_name', '')
        else:
            display_last = cen.get('last_name', '')
            display_first = cen.get('first_name', '')

        if inv and cen:
            # Both present — compare key fields
            inv_tier = norm_tier(inv.get('plan_election'))
            cen_tier = norm_tier(cen.get('plan_election'))

            if inv_tier and cen_tier and inv_tier != cen_tier:
                discrepancies.append({
                    'field': 'Coverage Tier',
                    'invoice_val': inv_tier,
                    'census_val': cen_tier,
                    'severity': 'error'
                })
                match_status = 'mismatch'

            # Name spelling differences
            inv_name = f"{inv.get('last_name','')} {inv.get('first_name','')}".strip().lower()
            cen_name = f"{cen.get('last_name','')} {cen.get('first_name','')}".strip().lower()
            if inv_name != cen_name:
                discrepancies.append({
                    'field': 'Name Spelling',
                    'invoice_val': f"{inv.get('last_name')}, {inv.get('first_name')}",
                    'census_val': f"{cen.get('last_name')}, {cen.get('first_name')}",
                    'severity': 'warning'
                })
                if match_status == 'matched':
                    match_status = 'warning'

            # Status check
            inv_status = (inv.get('emp_status') or '').upper()
            cen_status = (cen.get('emp_status') or '').upper()
            if inv_status and cen_status and inv_status != cen_status:
                discrepancies.append({
                    'field': 'Emp Status',
                    'invoice_val': inv_status,
                    'census_val': cen_status,
                    'severity': 'warning'
                })
                if match_status == 'matched':
                    match_status = 'warning'

        elif inv and not cen:
            match_status = 'invoice_only'
            discrepancies.append({
                'field': 'Missing from Census',
                'invoice_val': '✓ On Invoice',
                'census_val': '✗ Not in census',
                'severity': 'error'
            })

        elif cen and not inv:
            match_status = 'census_only'
            discrepancies.append({
                'field': 'Missing from Invoice',
                'invoice_val': '✗ Not on invoice',
                'census_val': '✓ In census',
                'severity': 'warning'
            })

        rows.append({
            'last_name': display_last,
            'first_name': display_first,
            'match_status': match_status,
            'discrepancy_count': len(discrepancies),
            'discrepancies': discrepancies,
            'invoice': {
                'plan_election': inv.get('plan_election') if inv else None,
                'coverage_code': inv.get('_coverage_code') if inv else None,
                'premium_total': inv.get('_premium_total') if inv else None,
                'source': inv.get('_source') if inv else None,
                'period': inv.get('_invoice_period') if inv else None,
                'carrier': inv.get('_carrier') if inv else None,
            } if inv else None,
            'census': {
                'plan_election': cen.get('plan_election') if cen else None,
                'relationship': cen.get('relationship') if cen else None,
                'gender': cen.get('gender') if cen else None,
                'dob': cen.get('dob') if cen else None,
                'state': cen.get('state') if cen else None,
                'hire_date': cen.get('hire_date') if cen else None,
                'source': cen.get('_source') if cen else None,
            } if cen else None,
        })

    # Sort: mismatches first, then invoice_only, then census_only, then matched
    order = {'mismatch': 0, 'invoice_only': 1, 'census_only': 2, 'warning': 3, 'matched': 4}
    rows.sort(key=lambda r: (order.get(r['match_status'], 5), r['last_name']))
    return rows


# ─────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    uploaded = []
    for f in request.files.getlist('files'):
        if not f.filename:
            continue
        save_path = app.config['UPLOAD_FOLDER'] / f.filename
        f.save(save_path)
        uploaded.append({"name": f.filename, "path": str(save_path),
                         "type": Path(f.filename).suffix.lower().strip('.')})
        session_data["files"].append({"name": f.filename, "path": str(save_path),
                                       "type": Path(f.filename).suffix.lower().strip('.')})
    return jsonify({"ok": True, "files": uploaded})

@app.route('/process', methods=['POST'])
def process_files():
    all_raw = []
    all_logs = []

    for file_info in session_data["files"]:
        path = file_info["path"]
        name = file_info["name"]
        ftype = file_info["type"]

        if ftype in ('xlsx', 'xls', 'csv'):
            records, log = parse_excel_csv(path, name)
            # Tag as census source
            for r in records:
                r.setdefault('_source_type', 'census')
        elif ftype == 'pdf':
            records, log = parse_pdf(path, name)
            # parse_pdf already tags _source_type='invoice' for carrier invoices
            for r in records:
                r.setdefault('_source_type', 'invoice' if r.get('_carrier') else 'census')
        else:
            records, log = [], [f"Unsupported file type: {ftype}"]

        all_raw.extend(records)
        all_logs.extend([f"[{name}] {l}" for l in log])

    # Normalize all records
    normalized = [normalize_record(r) for r in all_raw]

    # Merge duplicates across sources
    merged, merge_log = merge_records(normalized)

    # Validate
    issues = validate_records(merged)

    # Run reconcile on the pre-merge normalized records (keeps source types intact)
    reconcile_rows = build_reconcile(normalized)

    session_data["records"] = merged
    session_data["issues"] = issues
    session_data["merge_log"] = merge_log
    session_data["reconcile"] = reconcile_rows

    # Summary stats
    subscribers = [r for r in merged if r.get('relationship') == 'Subscriber']
    dependents = [r for r in merged if r.get('relationship') in ('Spouse', 'Child')]
    errors = [i for i in issues if i['severity'] == 'error']
    warnings = [i for i in issues if i['severity'] == 'warning']
    cross_source = [m for m in merge_log if m['action'] == 'merged']

    # Reconcile summary
    rec_matched    = sum(1 for r in reconcile_rows if r['match_status'] == 'matched')
    rec_mismatch   = sum(1 for r in reconcile_rows if r['match_status'] == 'mismatch')
    rec_inv_only   = sum(1 for r in reconcile_rows if r['match_status'] == 'invoice_only')
    rec_cen_only   = sum(1 for r in reconcile_rows if r['match_status'] == 'census_only')
    rec_warning    = sum(1 for r in reconcile_rows if r['match_status'] == 'warning')

    return jsonify({
        "ok": True,
        "stats": {
            "total": len(merged),
            "subscribers": len(subscribers),
            "dependents": len(dependents),
            "errors": len(errors),
            "warnings": len(warnings),
            "cross_source_merges": len(cross_source),
            "sources": len(session_data["files"])
        },
        "reconcile_summary": {
            "matched": rec_matched,
            "mismatch": rec_mismatch,
            "invoice_only": rec_inv_only,
            "census_only": rec_cen_only,
            "warning": rec_warning,
            "total": len(reconcile_rows),
            "has_data": len(reconcile_rows) > 0,
        },
        "issues": issues[:50],
        "merge_log": merge_log[:50],
        "logs": all_logs[:30],
        "detected_company": session_data.get("detected_company", ""),
        "company_name": session_data.get("company_name", ""),
    })

@app.route('/reconcile', methods=['GET'])
def get_reconcile():
    return jsonify({
        "rows": session_data.get("reconcile", []),
        "invoice_meta": session_data.get("invoice_meta", {}),
    })

@app.route('/company', methods=['POST'])
def set_company():
    data = request.json
    session_data["company_name"] = data.get("name", "")
    return jsonify({"ok": True})

@app.route('/records', methods=['GET'])
def get_records():
    return jsonify({"records": session_data["records"], "issues": session_data["issues"]})

@app.route('/records/update', methods=['POST'])
def update_record():
    data = request.json
    idx = data.get('index')
    field = data.get('field')
    value = data.get('value')
    if idx is not None and field and 0 <= idx < len(session_data["records"]):
        session_data["records"][idx][field] = value
        # Re-validate
        session_data["issues"] = validate_records(session_data["records"])
        return jsonify({"ok": True, "issues": len(session_data["issues"])})
    return jsonify({"ok": False})

@app.route('/export/<carrier>', methods=['GET'])
def export(carrier):
    records = session_data["records"]
    if not records:
        return jsonify({"error": "No records"}), 400

    export_dir = app.config['EXPORT_FOLDER']
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    if carrier == 'triton':
        company = session_data.get("company_name") or session_data.get("detected_company", "")
        wb = build_triton_census(records, company)
        fname = f"Triton_Master_Census_{timestamp}.xlsx"
    elif carrier in ('BCBS', 'Aetna', 'UHC', 'Cigna'):
        wb = build_carrier_file(records, carrier)
        fname = f"{carrier}_Census_{timestamp}.xlsx"
    else:
        return jsonify({"error": "Unknown carrier"}), 400

    fpath = export_dir / fname
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname)

@app.route('/export/all', methods=['POST'])
def export_all():
    """Export all selected carriers as a zip."""
    import zipfile
    carriers = request.json.get('carriers', ['triton'])
    records = session_data["records"]
    if not records:
        return jsonify({"error": "No records"}), 400

    export_dir = app.config['EXPORT_FOLDER']
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    zip_path = export_dir / f"CensusIQ_Export_{timestamp}.zip"

    with zipfile.ZipFile(zip_path, 'w') as zf:
        for carrier in carriers:
            if carrier == 'triton':
                wb = build_triton_census(records)
                fname = f"Triton_Master_Census_{timestamp}.xlsx"
            else:
                wb = build_carrier_file(records, carrier)
                fname = f"{carrier}_Census_{timestamp}.xlsx"
            fpath = export_dir / fname
            wb.save(fpath)
            zf.write(fpath, fname)

    return send_file(zip_path, as_attachment=True, download_name=f"CensusIQ_Export_{timestamp}.zip")

@app.route('/reset', methods=['POST'])
def reset():
    session_data["files"] = []
    session_data["records"] = []
    session_data["issues"] = []
    session_data["merge_log"] = []
    session_data["company_name"] = ""
    session_data["detected_company"] = ""
    session_data["invoice_meta"] = {}
    session_data["invoice_records"] = []
    session_data["census_records"] = []
    session_data["reconcile"] = []
    return jsonify({"ok": True})

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

if __name__ == '__main__':
    # Ensure folders exist
    Path('uploads').mkdir(exist_ok=True)
    Path('exports').mkdir(exist_ok=True)

    # Open browser after short delay
    def open_browser():
        time.sleep(1.2)
        webbrowser.open('http://localhost:5050')

    t = threading.Thread(target=open_browser, daemon=True)
    t.start()

    print("\n" + "="*50)
    print("  CensusIQ — Triton Benefits")
    print("  Running at http://localhost:5050")
    print("  Press Ctrl+C to stop")
    print("="*50 + "\n")

    app.run(port=5050, debug=False)
